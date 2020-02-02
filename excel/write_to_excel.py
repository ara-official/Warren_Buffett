from openpyxl import load_workbook 
from openpyxl import Workbook 
import openpyxl
import os

class Write_excel:
    def write_sample(self, 주식데이터, row_num, save_file_name, bPrint=False):
        if(os.path.isfile(save_file_name) == False):    #파일 없을 시 생성
            file_create = Workbook()
            file_create.save(save_file_name)
        
        write_wb = openpyxl.load_workbook(save_file_name)

        write_ws = write_wb.active
        write_ws.append(["0: 날짜1", "1: 시간", "2: 시가", "3: 고가", "4: 저가", "5: 종가", "6: 전일대비", "8: 거래량", "9: 거래대금", "10: 누적체결매도수량"])

        if bPrint == True:
            print('input Data : %s' % (주식데이터))
        for i in range (len(주식데이터)):
            write_ws.cell(row=row_num+2,column=i+1).value=주식데이터[i]

        
        write_wb.save(save_file_name)